from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import settings


PROJECT_SOURCE_KEYWORDS = ("project", "项目", "项目定制")
STANDARD_SOURCE_KEYWORDS = ("standard", "标准", "标准文件", "通用")
OTHER_SOURCE_KEYWORDS = ("other", "其他", "说明")
SOURCE_MATRIX_ENV = "TECHNICAL_APPENDIX_SOURCE_MATRIX_PATH"
DEFAULT_SOURCE_MATRIX_FILE_NAME = "technical_appendix_source_matrix.xlsx"
APPENDIX_CODE_RE = re.compile(
    r"附表\s*([A-Za-z]?\s*\.?\s*\d+(?:\.\d+)*)(?:\s*[-—~～至到]\s*([A-Za-z]?\s*\.?\s*\d+(?:\.\d+)*))?",
    re.IGNORECASE,
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\n", " / ")).strip()


def normalize_match_text(value: Any) -> str:
    return re.sub(r"[\s　,，、.。:：;；()（）\[\]【】{}<>《》\"'`·_\-—/\\|]+", "", clean_text(value).lower())


def normalize_appendix_code(value: Any) -> str:
    text = clean_text(value).upper()
    return re.sub(r"\s+", "", text).lstrip(".")


def appendix_code(value: Any) -> str:
    match = APPENDIX_CODE_RE.search(clean_text(value))
    return normalize_appendix_code(match.group(1)) if match else ""


def _appendix_code_parts(value: Any) -> tuple[str, tuple[int, ...]] | None:
    code = normalize_appendix_code(value)
    match = re.fullmatch(r"([A-Z]+)?\.?([0-9]+(?:\.[0-9]+)*)", code)
    if not match:
        return None
    prefix = match.group(1) or ""
    numbers = tuple(int(part) for part in match.group(2).split("."))
    return prefix, numbers


def appendix_rule_code_score(table_title: Any, rule_title: Any) -> float:
    table_code = appendix_code(table_title)
    if not table_code:
        return 0.0
    rule_match = APPENDIX_CODE_RE.search(clean_text(rule_title))
    if not rule_match:
        return 0.0
    start_code = normalize_appendix_code(rule_match.group(1))
    end_code = normalize_appendix_code(rule_match.group(2) or "")
    if not end_code:
        if table_code == start_code:
            return 0.96
        # 子编号覆盖：规则只写父级编号（如 F.2）时覆盖附表子编号（F.2.1/F.2.2）。
        # 分值低于精确命中与区间命中，更具体的规则（若有）仍然优先。
        table_parts = _appendix_code_parts(table_code)
        start_parts = _appendix_code_parts(start_code)
        if table_parts and start_parts:
            table_prefix, table_numbers = table_parts
            start_prefix, start_numbers = start_parts
            if not (table_prefix and start_prefix and table_prefix != start_prefix):
                if len(table_numbers) > len(start_numbers) and table_numbers[: len(start_numbers)] == start_numbers:
                    return 0.93
        return 0.0

    table_parts = _appendix_code_parts(table_code)
    start_parts = _appendix_code_parts(start_code)
    end_parts = _appendix_code_parts(end_code)
    if not table_parts or not start_parts or not end_parts:
        return 0.0
    table_prefix, table_numbers = table_parts
    start_prefix, start_numbers = start_parts
    end_prefix, end_numbers = end_parts
    if not end_prefix:
        end_prefix = start_prefix
    if table_prefix and start_prefix and table_prefix != start_prefix:
        return 0.0
    if table_prefix and end_prefix and table_prefix != end_prefix:
        return 0.0
    if start_numbers <= table_numbers <= end_numbers:
        return 0.94
    return 0.0


def source_terms(value: Any) -> list[str]:
    if isinstance(value, list):
        terms: list[str] = []
        seen: set[str] = set()
        for item in value:
            for term in source_terms(item):
                if term and term not in seen:
                    seen.add(term)
                    terms.append(term)
        return terms
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[&＆,，、;；/／\n]+", text)
    terms: list[str] = []
    seen: set[str] = set()
    for part in parts:
        item = clean_text(part).strip(" ：:")
        if not item or item in seen:
            continue
        seen.add(item)
        terms.append(item)
    return terms


def _header_kind(value: Any) -> str:
    text = normalize_match_text(value)
    if not text:
        return ""
    if "客户" in text:
        return "customer"
    if "表格" in text or "附表" in text:
        return "table"
    if any(normalize_match_text(keyword) in text for keyword in PROJECT_SOURCE_KEYWORDS):
        return "project"
    if any(normalize_match_text(keyword) in text for keyword in STANDARD_SOURCE_KEYWORDS):
        return "standard"
    if any(normalize_match_text(keyword) in text for keyword in OTHER_SOURCE_KEYWORDS):
        return "other"
    return ""


def _detect_header(values: list[list[Any]]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(values[:12]):
        mapping: dict[str, int] = {}
        for col, value in enumerate(row):
            kind = _header_kind(value)
            if kind and kind not in mapping:
                mapping[kind] = col
        if "customer" in mapping and "table" in mapping:
            return index, mapping
    return -1, {}


def parse_appendix_source_matrix(path: Path | str) -> dict[str, Any]:
    matrix_path = Path(path).expanduser()
    if not matrix_path.exists() or matrix_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {"schemaVersion": "technical-appendix-source-matrix-v1", "path": str(matrix_path), "rows": []}

    wb = load_workbook(matrix_path, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        values = [list(row) for row in ws.iter_rows(values_only=True)]
        header_index, mapping = _detect_header(values)
        if header_index < 0:
            continue
        for row_number, row in enumerate(values[header_index + 1 :], start=header_index + 2):
            def cell(kind: str) -> str:
                col = mapping.get(kind)
                return clean_text(row[col]) if col is not None and col < len(row) else ""

            customer = cell("customer")
            table_title = cell("table")
            if not customer or not table_title:
                continue
            project_sources = source_terms(cell("project"))
            standard_sources = source_terms(cell("standard"))
            other_sources = source_terms(cell("other"))
            if not project_sources and not standard_sources and not other_sources:
                continue
            rows.append(
                {
                    "id": f"{ws.title}!R{row_number}",
                    "sheet": ws.title,
                    "row": row_number,
                    "customer": customer,
                    "tableTitle": table_title,
                    "projectSources": project_sources,
                    "standardSources": standard_sources,
                    "otherSources": other_sources,
                }
            )

    return {
        "schemaVersion": "technical-appendix-source-matrix-v1",
        "path": str(matrix_path),
        "rows": rows,
    }


def resolve_appendix_source_matrix_path(project: dict[str, Any]) -> str:
    candidates = [
        project.get("technicalAppendixSourceMatrixPath"),
        project.get("appendixSourceMatrixPath"),
        (project.get("technicalAppendixSourceMatrix") or {}).get("path")
        if isinstance(project.get("technicalAppendixSourceMatrix"), dict)
        else "",
        os.getenv(SOURCE_MATRIX_ENV),
        settings.documents_dir / "_config" / DEFAULT_SOURCE_MATRIX_FILE_NAME,
    ]
    for candidate in candidates:
        text = clean_text(candidate)
        if not text:
            continue
        path = Path(text).expanduser()
        if path.exists() and path.is_file():
            return str(path)
    return ""


def load_appendix_source_matrix_for_project(project: dict[str, Any]) -> dict[str, Any]:
    path = resolve_appendix_source_matrix_path(project)
    if not path:
        return {"schemaVersion": "technical-appendix-source-matrix-v1", "path": "", "rows": []}
    return parse_appendix_source_matrix(path)


def table_title_match_score(left: Any, right: Any) -> float:
    code_score = appendix_rule_code_score(left, right)
    if code_score:
        return code_score
    left_norm = normalize_match_text(left)
    right_norm = normalize_match_text(right)
    if not left_norm or not right_norm:
        return 0.0
    if left_norm == right_norm:
        return 1.0
    if left_norm in right_norm or right_norm in left_norm:
        return 0.88
    shared = len(set(left_norm) & set(right_norm))
    total = len(set(left_norm) | set(right_norm))
    return shared / total if total else 0.0


def customer_match_score(project_customer: Any, rule_customer: Any) -> float:
    project_norm = normalize_match_text(project_customer)
    rule_norm = normalize_match_text(rule_customer)
    if not project_norm or not rule_norm:
        return 0.0
    if project_norm == rule_norm:
        return 1.0
    if project_norm in rule_norm or rule_norm in project_norm:
        return 0.9
    return 0.0


def find_appendix_source_rule(
    matrix: dict[str, Any],
    *,
    customer_name: Any,
    table_title: Any,
) -> dict[str, Any]:
    rows = matrix.get("rows") if isinstance(matrix.get("rows"), list) else []
    best: tuple[float, dict[str, Any]] | None = None
    for row in rows:
        if not isinstance(row, dict):
            continue
        table_score = table_title_match_score(table_title, row.get("tableTitle"))
        if table_score < 0.82:
            continue
        customer_score = customer_match_score(customer_name, row.get("customer"))
        if customer_name and customer_score <= 0:
            continue
        score = table_score * 10 + customer_score * 3
        if best is None or score > best[0]:
            best = (score, row)
    if best is None:
        return {}
    return dict(best[1])


# ---------------------------------------------------------------------------
# 已生成 gap plan 的矩阵回写
#
# 以下三个函数镜像 bid-tech-gap-planner skill（run_from_manifest.py）里的
# matrix_material_score / source_routing_payload / 推荐素材裁剪逻辑，
# 用于「素材匹配完成后才上传规则」的场景：不必重跑整个缺口识别，
# 直接把矩阵应用到现在 plan 的附表任务上。权重与 skill 保持一致，
# 避免同一项目两次路由结果排序漂移。
# ---------------------------------------------------------------------------

MATERIAL_TEXT_KEYS = ("name", "path", "docx", "cleanedPath", "folderPath", "cleanedFileName", "matchReason")


def _material_text(material: dict[str, Any]) -> str:
    return " ".join(str(material.get(key) or "") for key in MATERIAL_TEXT_KEYS)


def _tier_key(material: dict[str, Any]) -> str:
    return normalize_match_text(material.get("materialTier") or material.get("materialScope") or "")


def _tier_is_project(tier: str) -> bool:
    return "project" in tier or "项目" in tier


def _tier_is_standard(tier: str) -> bool:
    return "standard" in tier or "标准" in tier or "通用" in tier


def matrix_material_score(material: dict[str, Any], rule: dict[str, Any]) -> tuple[float, list[str]]:
    """规则来源词对单条素材打分（与 skill 的 matrix_material_score 同权重）。"""
    terms = {
        "project": source_terms(rule.get("projectSources")),
        "standard": source_terms(rule.get("standardSources")),
    }
    if not any(terms.values()):
        return 0.0, []
    text = normalize_match_text(_material_text(material))
    tier = _tier_key(material)
    score = 0.0
    reasons: list[str] = []
    for scope, scope_terms in (("project", terms["project"]), ("standard", terms["standard"])):
        scope_hit = _tier_is_project(tier) if scope == "project" else _tier_is_standard(tier)
        for term in scope_terms:
            term_key = normalize_match_text(term)
            if not term_key:
                continue
            if term_key in text:
                score += 420 if scope_hit else 260
                reasons.append(f"{scope} 来源规定命中：{term}")
            elif any(
                part_key and len(part_key) >= 2 and part_key in text
                for part in source_terms(term)
                for part_key in [normalize_match_text(part)]
            ):
                score += 180 if scope_hit else 120
                reasons.append(f"{scope} 来源规定部分命中：{term}")
    if score and _tier_is_project(tier):
        score += 30
    elif score and _tier_is_standard(tier):
        score += 20
    return score, reasons[:6]


def build_source_routing_payload(rule: dict[str, Any], matched_materials: list[dict[str, Any]]) -> dict[str, Any]:
    """附表任务/目录项上的 sourceRouting 载荷（与 skill 的 source_routing_payload 同构）。"""
    if not rule:
        return {}
    project_sources = source_terms(rule.get("projectSources"))
    standard_sources = source_terms(rule.get("standardSources"))
    other_sources = source_terms(rule.get("otherSources"))
    matched = [
        {
            "id": str(material.get("id") or material.get("materialId") or ""),
            "name": str(material.get("name") or material.get("cleanedFileName") or ""),
            "folderPath": str(material.get("folderPath") or ""),
            "materialTier": str(material.get("materialTier") or ""),
            "matchReason": str(material.get("matchReason") or ""),
        }
        for material in matched_materials[:8]
        if material.get("sourceRouting")
    ]
    manual_terms = [term for term in other_sources if any(token in term for token in ("人工", "收集", "项目定制收集"))]
    tender_terms = [term for term in other_sources if any(token in term for token in ("招标", "响应招标"))]
    status = "matched" if matched else ("manual_required" if manual_terms else ("tender_parse_fields" if tender_terms else "missing_source"))
    return {
        "status": status,
        "source": "appendix_source_matrix",
        "ruleId": str(rule.get("id") or ""),
        "customer": str(rule.get("customer") or ""),
        "tableTitle": str(rule.get("tableTitle") or ""),
        "projectSources": project_sources,
        "standardSources": standard_sources,
        "otherSources": other_sources,
        "matchedMaterials": matched,
        "manualRequired": bool(manual_terms),
        "useTenderParseFields": bool(tender_terms),
    }


def route_materials_for_rule(rule: dict[str, Any], materials: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按规则行给素材池打分、过滤、排序（与 skill 的 recommended_materials_for_appendix 同语义）。"""
    ranked: list[tuple[float, dict[str, Any]]] = []
    for material in materials:
        if not isinstance(material, dict):
            continue
        score, reasons = matrix_material_score(material, rule)
        if not score:
            continue
        item = dict(material)
        item["matchReason"] = "；".join([*(reasons or []), str(item.get("matchReason") or "")]).strip("；")
        item["sourceRouting"] = {
            "source": "appendix_source_matrix",
            "ruleId": str(rule.get("id") or ""),
            "reasons": reasons,
        }
        ranked.append((score, item))
    ranked.sort(key=lambda entry: entry[0], reverse=True)
    return [item for _, item in ranked]


def apply_appendix_source_matrix_to_plan(
    plan: dict[str, Any],
    matrix: dict[str, Any],
    *,
    customer_name: str,
    materials: list[dict[str, Any]],
) -> dict[str, int]:
    """把来源矩阵回写到已生成的 gap plan：新矩阵完整覆盖旧矩阵路由。

    先清除上一版矩阵产生的任务/目录项路由，再应用新规则；非矩阵来源（如甲方
    已填附表）和非矩阵推荐素材保持不变。目录项同步首个命中任务的路由与素材，
    与 skill 产出结构一致，前端弹窗与 AI 填写 manifest 直接可用。

    stats 里 clearedItems/clearedTasks 记录本次清除的旧矩阵路由数量：调用方
    据此判断是否需要持久化——新版规则零命中时 routedItems 为 0，但清除本身
    也是改动，必须落库，否则旧路由会在重新读取时复活（R10-B09-03）。
    """
    stats = {
        "routedItems": 0,
        "matchedTasks": 0,
        "manualRequired": 0,
        "tenderFields": 0,
        "missingSource": 0,
        "clearedItems": 0,
        "clearedTasks": 0,
    }
    rows = matrix.get("rows") if isinstance(matrix.get("rows"), list) else []
    items = plan.get("items") if isinstance(plan.get("items"), list) else []
    if not rows or not items:
        return stats
    for item in items:
        if not isinstance(item, dict):
            continue
        item_routing = item.get("sourceRouting") if isinstance(item.get("sourceRouting"), dict) else {}
        if item_routing.get("source") == "appendix_source_matrix":
            item.pop("sourceRouting", None)
            item.pop("sourceRoutedMaterials", None)
            stats["clearedItems"] += 1
        appendix_tasks = item.get("appendixTasks") if isinstance(item.get("appendixTasks"), list) else []
        item_payload: dict[str, Any] = {}
        item_materials: list[dict[str, Any]] = []
        for task in appendix_tasks:
            if not isinstance(task, dict):
                continue
            task_routing = task.get("sourceRouting") if isinstance(task.get("sourceRouting"), dict) else {}
            if task_routing.get("source") == "appendix_source_matrix":
                task.pop("sourceRouting", None)
                stats["clearedTasks"] += 1
                recommended = task.get("recommendedMaterials")
                if isinstance(recommended, list):
                    task["recommendedMaterials"] = [
                        material
                        for material in recommended
                        if not (
                            isinstance(material, dict)
                            and isinstance(material.get("sourceRouting"), dict)
                            and material["sourceRouting"].get("source") == "appendix_source_matrix"
                        )
                    ]
            elif task_routing.get("source"):
                continue
            title = str(task.get("title") or task.get("id") or "")
            rule = find_appendix_source_rule(matrix, customer_name=customer_name, table_title=title)
            if not rule:
                continue
            routed = route_materials_for_rule(rule, materials)
            payload = build_source_routing_payload(rule, routed)
            task["sourceRouting"] = payload
            task["recommendedMaterials"] = [
                {**dict(material), "usage": "table_source"} for material in routed[:5]
            ]
            if not item_payload:
                item_payload = payload
                item_materials = task["recommendedMaterials"]
            status = str(payload.get("status") or "")
            if status == "matched":
                stats["matchedTasks"] += 1
            elif status == "manual_required":
                stats["manualRequired"] += 1
            elif status == "tender_parse_fields":
                stats["tenderFields"] += 1
            else:
                stats["missingSource"] += 1
        if item_payload:
            item["sourceRouting"] = item_payload
            item["sourceRoutedMaterials"] = item_materials
            stats["routedItems"] += 1
    return stats
