from __future__ import annotations

import io
import re
from pathlib import PurePosixPath
from typing import Any

from openpyxl import load_workbook


MODEL_PATTERN = re.compile(
    r"\b(?:E?W|SE)\d+(?:\.\d+)?-\d{3}(?:[A-Za-z0-9_（）()\-]*|(?:上置|下置|海外版|碳叶片))*",
    re.IGNORECASE,
)
EXPLICIT_MODEL_PATTERN = re.compile(r"^(?:E?W|SE)\d+(?:\.\d+)?-\d{3}", re.IGNORECASE)
NOISE_PATTERN = re.compile(
    r"^(?:CGC|CQC|CN|GB|IEC|ISO|HB|S)\d+|^\d{6,}$|^20\d{2}|^[A-Z]{1,4}-[A-Z]{2,}\d+",
    re.IGNORECASE,
)
LAYOUT_WORDS = ("上置", "下置")


def normalize_project_turbine_model(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, str):
        model = value.strip()
        raw: dict[str, Any] = {"model": model} if model else {}
    elif isinstance(value, dict):
        raw = dict(value)
        model = str(
            raw.get("model")
            or raw.get("turbineModel")
            or raw.get("machineModel")
            or raw.get("name")
            or raw.get("label")
            or ""
        ).strip()
        if not model:
            return {}
        raw["model"] = model
    else:
        return {}

    aliases = model_aliases(str(raw.get("model") or ""))
    for alias in raw.get("aliases") or raw.get("turbineAliases") or []:
        text = str(alias or "").strip()
        if text and text not in aliases:
            aliases.append(text)
    return {
        "model": str(raw.get("model") or "").strip(),
        "platform": str(raw.get("platform") or raw.get("turbinePlatform") or "").strip(),
        "layout": str(raw.get("layout") or raw.get("transformerLayout") or "").strip(),
        "ratedPowerKw": _number_or_empty(raw.get("ratedPowerKw") or raw.get("rated_power_kw")),
        "rotorDiameterM": _number_or_empty(raw.get("rotorDiameterM") or raw.get("rotor_diameter_m")),
        "status": str(raw.get("status") or "manual").strip() or "manual",
        "statusLabel": str(raw.get("statusLabel") or raw.get("statusText") or "").strip(),
        "source": str(raw.get("source") or "manual").strip() or "manual",
        "sourceFileId": str(raw.get("sourceFileId") or "").strip(),
        "sourceFileName": str(raw.get("sourceFileName") or "").strip(),
        "aliases": aliases,
    }


def project_turbine_model(project: dict[str, Any]) -> dict[str, Any]:
    return normalize_project_turbine_model(
        project.get("turbineModel")
        or project.get("selectedTurbineModel")
        or project.get("machineModel")
        or project.get("turbineModelLabel")
    )


def model_aliases(model: str) -> list[str]:
    text = _clean_model_text(model)
    if not text:
        return []
    variants = [text, text.replace("_", ""), text.replace("_", "-")]
    for word in LAYOUT_WORDS:
        if word in text:
            variants.append(text.replace(word, ""))
    for item in list(variants):
        normalized = item.strip("_- ")
        if normalized.upper().startswith("EW"):
            variants.append(normalized[1:])
        elif normalized.upper().startswith("W"):
            variants.append(f"E{normalized}")
    output: list[str] = []
    for item in variants:
        clean = item.strip("_- ")
        if clean and clean not in output:
            output.append(clean)
    return output


def extract_turbine_model_options_from_xlsx_bytes(
    data: bytes,
    *,
    source_file_id: str = "",
    source_file_name: str = "",
    folder_path: str = "",
) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    options: list[dict[str, Any]] = []
    seen: set[str] = set()
    for sheet in workbook.worksheets:
        if sheet.max_row < 6:
            continue
        for column in range(1, sheet.max_column + 1):
            raw_model = _cell_text(sheet.cell(row=5, column=column).value)
            model = _clean_model_text(raw_model)
            if not is_valid_turbine_model(model):
                continue
            platform = _cell_text(sheet.cell(row=4, column=column).value)
            rated_power = _number_or_empty(sheet.cell(row=6, column=column).value)
            rotor_diameter = _number_or_empty(sheet.cell(row=8, column=column).value)
            layout = _infer_layout(
                model,
                _cell_text(sheet.cell(row=2, column=column).value),
            )
            status_text = " ".join(
                part
                for part in (
                    _cell_text(sheet.cell(row=1, column=column).value),
                    _cell_text(sheet.cell(row=2, column=column).value),
                )
                if part
            )
            status = classify_model_status(status_text)
            key = f"{model}|{platform}|{layout}"
            if key in seen:
                continue
            seen.add(key)
            options.append(
                {
                    "id": _option_id(model, platform, layout),
                    "model": model,
                    "label": model,
                    "platform": platform,
                    "layout": layout,
                    "ratedPowerKw": rated_power,
                    "rotorDiameterM": rotor_diameter,
                    "status": status,
                    "statusLabel": status_label(status, status_text),
                    "source": "parameter_sheet",
                    "sourceFileId": source_file_id,
                    "sourceFileName": source_file_name,
                    "folderPath": folder_path,
                    "aliases": model_aliases(model),
                    "evidence": [
                        {
                            "sheet": sheet.title,
                            "row": 5,
                            "column": column,
                            "sourceFileId": source_file_id,
                            "sourceFileName": source_file_name,
                        }
                    ],
                }
            )
    return sorted(options, key=_sort_key)


def is_valid_turbine_model(value: str) -> bool:
    text = _clean_model_text(value)
    if not text:
        return False
    if NOISE_PATTERN.search(text):
        return False
    if not EXPLICIT_MODEL_PATTERN.match(text):
        return False
    return "-" in text and any(char.isdigit() for char in text)


def classify_model_status(text: str) -> str:
    raw = str(text or "")
    if "废弃" in raw:
        return "deprecated"
    if "未立项" in raw:
        return "not_approved"
    if "研发" in raw:
        return "research"
    return "active"


def status_label(status: str, raw_text: str = "") -> str:
    if raw_text.strip():
        return raw_text.strip()
    return {
        "active": "可投标",
        "research": "研发中",
        "not_approved": "投标支持未立项",
        "deprecated": "废弃",
        "manual": "人工指定",
    }.get(status, status)


def material_model_fit(material: dict[str, Any], turbine_model: dict[str, Any]) -> str:
    selected = normalize_project_turbine_model(turbine_model)
    if not selected:
        return "generic"
    text = " ".join(
        str(material.get(key) or "")
        for key in (
            "name",
            "fileName",
            "folderPath",
            "sourceRelativePath",
            "sourceRootFolder",
            "turbineModel",
            "turbineModelLabel",
        )
    )
    tokens = [match.group(0) for match in MODEL_PATTERN.finditer(text)]
    if not tokens:
        return "generic"
    aliases = {alias.upper().replace("_", "") for alias in selected.get("aliases") or []}
    aliases.add(str(selected.get("model") or "").upper().replace("_", ""))
    for token in tokens:
        normalized = token.upper().replace("_", "")
        if normalized in aliases:
            return "match"
        for alias in aliases:
            if alias and (alias in normalized or normalized in alias):
                return "match"
    return "conflict"


def turbine_model_from_material_name(name: str, *, source_file_id: str = "", folder_path: str = "") -> dict[str, Any] | None:
    match = MODEL_PATTERN.search(str(name or ""))
    if not match:
        return None
    model = _clean_model_text(match.group(0))
    if not is_valid_turbine_model(model):
        return None
    return normalize_project_turbine_model(
        {
            "model": model,
            "layout": _infer_layout(model, ""),
            "source": "material_name",
            "sourceFileId": source_file_id,
            "sourceFileName": PurePosixPath(str(name or "")).name,
            "folderPath": folder_path,
            "status": "material_hint",
            "statusLabel": "来自素材文件名",
        }
    )


def _sort_key(item: dict[str, Any]) -> tuple[int, float, str]:
    status_order = {"active": 0, "research": 1, "not_approved": 2, "deprecated": 3}
    power = item.get("ratedPowerKw")
    try:
        power_value = -float(power or 0)
    except (TypeError, ValueError):
        power_value = 0.0
    return (status_order.get(str(item.get("status") or ""), 9), power_value, str(item.get("model") or ""))


def _clean_model_text(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", "", text)
    return text.strip("；;，,。")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number_or_empty(value: Any) -> int | float | str:
    if value is None or value == "":
        return ""
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return ""
    return int(number) if number.is_integer() else number


def _infer_layout(model: str, hint: str) -> str:
    for source in (model, hint):
        for word in LAYOUT_WORDS:
            if word in str(source or ""):
                return f"变压器{word}"
    return str(hint or "").strip()


def _option_id(model: str, platform: str, layout: str) -> str:
    raw = "-".join(part for part in (model, platform, layout) if part)
    return re.sub(r"[^A-Za-z0-9\u4e00-\u9fff_.-]+", "-", raw).strip("-") or model
