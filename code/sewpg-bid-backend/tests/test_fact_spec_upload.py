from __future__ import annotations

"""技术标填表规则清单上传（POST /api/settings/technical-fact-specs）与 spec override 加载测试。"""

import json
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.core.config import settings
from app.main import app
from app.services import technical_fact_field_specs as specs_module
from app.services.store import store
from app.services.technical_fact_spec_import import EXPECTED_HEADER, FactSpecImportError, import_specs


@pytest.fixture()
def override_path(tmp_path, monkeypatch) -> Path:
    path = tmp_path / "documents" / "technical_fact_field_specs.override.json"
    monkeypatch.setattr(settings, "documents_dir", tmp_path / "documents")
    monkeypatch.setattr(settings, "fact_specs_override_path", path)
    settings.ensure_dirs()
    specs_module.clear_specs_cache()
    yield path
    specs_module.clear_specs_cache()


@pytest.fixture()
def client(override_path):
    store.reset_for_tests()
    with TestClient(app, base_url="http://127.0.0.1:8000") as test_client:
        login = test_client.post(
            "/api/auth/login", json={"email": "admin@sewpg.com", "password": "123456"}
        )
        assert login.status_code == 200
        test_client.headers["Authorization"] = f"Bearer {login.json()['token']}"
        yield test_client


def _build_xlsx(path: Path, rows: int = 5, header: list[str] | None = None, note: str = "说明") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header if header is not None else EXPECTED_HEADER)
    for index in range(1, rows + 1):
        ws.append(
            [
                index,
                "招标文件-技术规范书",
                "第一章 1.1",
                f"上传测试字段{index}",
                note,
                "",
                "招标文件/技术规范书",
            ]
        )
    wb.save(path)
    return path


def _upload(client: TestClient, path: Path, filename: str = "清单.xlsx"):
    with path.open("rb") as handle:
        return client.post(
            "/api/settings/technical-fact-specs",
            files={"file": (filename, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )


def test_upload_valid_xlsx_writes_override_and_matches_contract(client, override_path, tmp_path) -> None:
    xlsx_path = _build_xlsx(tmp_path / "清单.xlsx", rows=6)
    response = _upload(client, xlsx_path)

    assert response.status_code == 200
    payload = response.json()
    assert payload == {
        "specTotal": 6,
        "fillableTotal": 6,
        "needsConfirmation": 0,
        "template": 0,
        "override": True,
    }
    assert override_path.is_file()

    specs = specs_module.load_specs()
    assert len(specs) == 6
    assert specs[0]["label"] == "上传测试字段1"
    assert {spec["sourceKind"] for spec in specs} == {"tender"}


def test_upload_needs_confirmation_and_template_rows_counted(client, override_path, tmp_path) -> None:
    xlsx_path = _build_xlsx(tmp_path / "清单.xlsx", rows=3, note="需确认：以招标文件为准")
    response = _upload(client, xlsx_path)

    assert response.status_code == 200
    payload = response.json()
    assert payload["specTotal"] == 3
    assert payload["needsConfirmation"] == 3
    assert payload["fillableTotal"] == 3


def test_upload_rejects_non_xlsx_extension(client, override_path, tmp_path) -> None:
    bad_path = tmp_path / "清单.txt"
    bad_path.write_text("not an xlsx", encoding="utf-8")
    response = _upload(client, bad_path, filename="清单.txt")

    assert response.status_code == 400
    assert response.json()["detail"]
    assert not override_path.exists()


def test_upload_rejects_wrong_header_and_keeps_repo_default(client, override_path, tmp_path) -> None:
    xlsx_path = _build_xlsx(tmp_path / "清单.xlsx", rows=2, header=["A", "B", "C", "D", "E", "F", "G"])
    response = _upload(client, xlsx_path)

    assert response.status_code == 400
    assert "表头" in response.json()["detail"]
    assert not override_path.exists()
    # override 未落盘，load_specs 仍读仓库默认 148 条
    assert len(specs_module.load_specs()) == 148


def test_upload_rejects_empty_workbook(client, override_path, tmp_path) -> None:
    wb = openpyxl.Workbook()
    empty_path = tmp_path / "空.xlsx"
    wb.save(empty_path)
    response = _upload(client, empty_path)

    assert response.status_code == 400
    assert not override_path.exists()


def test_override_takes_priority_and_mtime_change_reloads(client, override_path, tmp_path) -> None:
    first = _build_xlsx(tmp_path / "第一版.xlsx", rows=4)
    assert _upload(client, first).status_code == 200
    assert len(specs_module.load_specs()) == 4

    # 绕过路由直接改写 override 文件：mtime 变化后 load_specs 自动重读（无需手工清缓存）
    second_specs = import_specs(_build_xlsx(tmp_path / "第二版.xlsx", rows=7))
    override_path.write_text(json.dumps(second_specs, ensure_ascii=False), encoding="utf-8")
    reloaded = specs_module.load_specs()
    assert len(reloaded) == 7
    assert reloaded[0]["label"] == "上传测试字段1"


def test_deleting_override_falls_back_to_repo_default(client, override_path, tmp_path) -> None:
    xlsx_path = _build_xlsx(tmp_path / "清单.xlsx", rows=3)
    assert _upload(client, xlsx_path).status_code == 200
    assert len(specs_module.load_specs()) == 3

    override_path.unlink()
    fallback = specs_module.load_specs()
    assert len(fallback) == 148
    assert len(specs_module.fillable_specs()) == 128


def test_corrupt_override_falls_back_to_repo_default(override_path) -> None:
    override_path.write_text("{broken json", encoding="utf-8")
    assert len(specs_module.load_specs()) == 148


def test_import_specs_writes_output_path(tmp_path) -> None:
    xlsx_path = _build_xlsx(tmp_path / "清单.xlsx", rows=2)
    output_path = tmp_path / "out" / "specs.json"
    specs = import_specs(xlsx_path, output_path=output_path)

    assert output_path.is_file()
    assert json.loads(output_path.read_text(encoding="utf-8")) == specs


def test_import_specs_rejects_bad_seq(tmp_path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EXPECTED_HEADER)
    ws.append(["不是数字", "招标文件", "第一章", "字段A", "", "", "招标文件/x"])
    bad_path = tmp_path / "坏序号.xlsx"
    wb.save(bad_path)

    with pytest.raises(FactSpecImportError, match="序号无效"):
        import_specs(bad_path)
